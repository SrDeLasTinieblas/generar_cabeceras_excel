import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border

# Abre el archivo Excel
archivo_excel = openpyxl.load_workbook('C:\\Users\\SrDeLasTinieblas\\Downloads\\Plantilla_Asiento.xlsx')

hoja = archivo_excel.active


# Recorre todas las celdas en la hoja
for fila in hoja.iter_rows():
    for celda in fila:
        # Verifica si la celda está fusionada
        if celda.coordinate in hoja.merged_cells:
            print(f'Celda fusionada: {hoja.merged_cells}')

        # Obtiene el valor de la celda
        valor = celda.value
        if valor:
            print(f'Valor de la celda {celda.coordinate}: {valor}')

        # Obtiene el estilo de la celda
        estilo = celda.font
        alineacion = celda.alignment
        fondo = celda.fill
        borde = celda.border

        # Aquí puedes imprimir o procesar los estilos según tus necesidades
        
        print(f'Estilo de la celda {celda.coordinate}:')
        print(f'Fuente: {estilo.name}, Tamaño: {estilo.size}, Color: {estilo.color}')
        print(f'Alineación: Horizontal: {alineacion.horizontal}, Vertical: {alineacion.vertical}')
        print(f'Color de fondo: {fondo.fgColor}')
        print(f'Borde: Top: {borde.top.style}, Left: {borde.left.style}, Right: {borde.right.style}, Bottom: {borde.bottom.style}')
        
        print(f'-------------------------------------------------------------------------------------------')

archivo_excel.close()
