import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border

# Abre el archivo Excel
archivo_excel = load_workbook('C:\\Users\\SrDeLasTinieblas\\Downloads\\Plantilla_Asiento.xlsx')

hoja = archivo_excel.active

# Celdas fusionadas
celdas_fusionadas = [range_.coord.split(":") for range_ in hoja.merged_cells.ranges]

codigo_csharp = ''

def get_column_number(columna):
    """
    Convierte una letra de columna en su equivalente numérico.
    Por ejemplo, 'A' se convierte en 1, 'B' en 2, 'Z' en 26, 'AA' en 27, 'AB' en 28, etc.
    """
    num = 0
    for c in columna:
        num = num * 26 + ord(c) - ord('A') + 1
    return num


for coord_inicial, coord_final in celdas_fusionadas:
    columna_inicial, fila_inicial = openpyxl.utils.cell.coordinate_from_string(coord_inicial)
    columna_final, fila_final = openpyxl.utils.cell.coordinate_from_string(coord_final)

    # Obtiene el número de la columna
    numero_columna = get_column_number(columna_inicial)
    # Agrega código para aplicar estilos en C#
    codigo_csharp += f'using (ExcelRange r = worksheet.Cells["{columna_inicial}{fila_inicial}:{columna_final}{fila_final}"])\n'
    codigo_csharp += '{\n'
    codigo_csharp += '    r.Merge = true;\n'
    codigo_csharp += '    r.Style.Font.SetFromFont(new Font("Calibri", 11));\n'
    codigo_csharp += '    r.Style.Font.Color.SetColor(Color.Black);\n'
    codigo_csharp += '    r.Style.HorizontalAlignment = ExcelHorizontalAlignment.Center;\n'
    codigo_csharp += '    r.Style.WrapText = true;\n'
    codigo_csharp += '    r.Style.Fill.PatternType = ExcelFillStyle.Solid;\n'
    codigo_csharp += '    r.Style.Fill.BackgroundColor.SetColor(Color.White);\n'
    codigo_csharp += '    r.Style.Font.Bold = false;\n'
    codigo_csharp += '    r.Style.Border.Top.Style = ExcelBorderStyle.Thin;\n'
    codigo_csharp += '    r.Style.Border.Left.Style = ExcelBorderStyle.Thin;\n'
    codigo_csharp += '    r.Style.Border.Right.Style = ExcelBorderStyle.Thin;\n'
    codigo_csharp += '    r.Style.Border.Bottom.Style = ExcelBorderStyle.Thin;\n'
    codigo_csharp += '}\n'
    codigo_csharp += f'worksheet.Column({numero_columna}).AutoFit();\n\n'
    
# Recorre todas las celdas fusionadas
for celdas_rango in hoja.merged_cells.ranges:
    coord_inicial = celdas_rango.coord.split(':')[0]
    coord_final = celdas_rango.coord.split(':')[1]

    letra_columna = coord_inicial[0]  # Obtiene la letra de la columna
    numero_fila_inicial = int(coord_inicial[1:])  # Obtiene el número de la fila inicial

    # Convierte la letra de la columna en el número correspondiente
    numero_columna = get_column_number(letra_columna)

    # Obtiene el valor de la celda fusionada
    valor_celda = hoja[coord_inicial].value

    if valor_celda:
        # Agrega código para establecer el valor de la celda en C#
        codigo_csharp += f'worksheet.Cells["{letra_columna}{numero_fila_inicial}"].Value = "{valor_celda}";\n'

# Ruta del archivo de texto donde deseas guardar el código
ruta_archivo = 'codigo_csharp.txt'  # Reemplaza 'ruta/del/archivo' con la ubicación deseada

# Guarda el código C# en un archivo de texto
with open(ruta_archivo, 'w') as archivo_txt:
    archivo_txt.write(codigo_csharp)

archivo_excel.close()
