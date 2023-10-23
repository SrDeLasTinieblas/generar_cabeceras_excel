import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.utils import coordinate_to_tuple, get_column_letter

#ruta_excel = 'C:\\Users\\SrDeLasTinieblas\\Downloads\\comprobacion_saldo.xlsx'
ruta_excel = 'cabecera_planilla.xlsx'

workbook = openpyxl.load_workbook(ruta_excel)
hoja = workbook.active

codigo_csharp = ""  # Define la variable código C# como una cadena vacía

celdas_no_fusionadas = []
celdas_fusionadas = []

def obtener_celdas_fusionadas(hoja):
    try:
        celdas_fusionadas = hoja.merged_cells.ranges
        rangos_fusionados = [str(rango) for rango in celdas_fusionadas]
        return rangos_fusionados
    except Exception as e:
        print(f"Error al procesar el archivo de Excel: {str(e)}")
        return []

def obtener_celdas_no_fusionadas(hoja, celda_inicio, celda_fin):
    try:
        #celdas_no_fusionadas = []
        fila_inicio, columna_inicio = coordinate_to_tuple(celda_inicio)
        fila_fin, columna_fin = coordinate_to_tuple(celda_fin)
        for fila in hoja.iter_rows(min_row=fila_inicio, max_row=fila_fin, min_col=columna_inicio, max_col=columna_fin):
            for celda in fila:
                if not any(celda.coordinate in rango for rango in hoja.merged_cells.ranges):
                    celdas_no_fusionadas.append(celda.coordinate)
        return celdas_no_fusionadas
    except Exception as e:
        print(f"Error al procesar el archivo de Excel: {str(e)}")
        return []

def generar_codigo_csharp(hoja, celdas_fusionadas, celdas_no_fusionadas):
    codigo_csharp = ""
    try:
        for rango in celdas_fusionadas:
            inicio, fin = rango.split(":")
            estilo_primera_celda = hoja[inicio].font
            nombre_fuente = estilo_primera_celda.name
            tamaño_fuente = int(estilo_primera_celda.size)
            codigo_csharp += f'using (ExcelRange r = worksheet.Cells["{rango}"])\n'
            codigo_csharp += '{\n'
            codigo_csharp += '    r.Merge = true;\n'
            codigo_csharp += f'    r.Style.Font.SetFromFont("{nombre_fuente}", {tamaño_fuente});\n'
            #codigo_csharp += f'    r.Style.Font.SetFromFont(new Font("{nombre_fuente}", {tamaño_fuente}));\n'
            codigo_csharp += '    r.Style.Font.Color.SetColor(Color.Black);\n'
            codigo_csharp += '    r.Style.HorizontalAlignment = ExcelHorizontalAlignment.Center;\n'
            codigo_csharp += '    r.Style.WrapText = true;\n'
            codigo_csharp += '    r.Style.Fill.PatternType = ExcelFillStyle.Solid;\n'
            codigo_csharp += '    r.Style.Fill.BackgroundColor.SetColor(Color.White);\n'
            codigo_csharp += '    r.Style.Font.Bold = false;\n'
            codigo_csharp += '    r.Style.Border.Top.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Left.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Right.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Bottom.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '}\n'
            columna = get_column_number(inicio)
            codigo_csharp += f'worksheet.Column({columna}).AutoFit();\n\n'
        for coordenada in celdas_no_fusionadas:
            estilo_celda = hoja[coordenada].font
            nombre_fuente = estilo_celda.name
            tamaño_fuente = int(estilo_celda.size)
            codigo_csharp += f'using (ExcelRange r = worksheet.Cells["{coordenada}"])\n'
            codigo_csharp += '{\n'
            codigo_csharp += f'    r.Style.Font.SetFromFont("{nombre_fuente}", {tamaño_fuente});\n'
            #codigo_csharp += f'    r.Style.Font.SetFromFont(new Font("{nombre_fuente}", {tamaño_fuente}));\n'
            codigo_csharp += '    r.Style.Font.Color.SetColor(Color.Black);\n'
            codigo_csharp += '    r.Style.HorizontalAlignment = ExcelHorizontalAlignment.Center;\n'
            codigo_csharp += '    r.Style.WrapText = true;\n'
            codigo_csharp += '    r.Style.Fill.PatternType = ExcelFillStyle.Solid;\n'
            codigo_csharp += '    r.Style.Fill.BackgroundColor.SetColor(Color.White);\n'
            codigo_csharp += '    r.Style.Font.Bold = false;\n'
            codigo_csharp += '    r.Style.Border.Top.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Left.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Right.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Bottom.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '}\n'
            columna = get_column_number(coordenada[:-1])
            codigo_csharp += f'worksheet.Column({columna}).AutoFit();\n'
        return codigo_csharp
    except Exception as e:
        print(f"Error al procesar el archivo de Excel: {str(e)}")
        return []
    
def generar_codigo_csharp_valores(hoja, celdas_fusionadas, celdas_no_fusionadas):
    codigo_csharp = ""  # Define la variable código C# como una cadena vacía
    # Recorre todas las celdas fusionadas
    print("Recorre todas las celdas fusionadas")
    for celdas_rango in hoja.merged_cells.ranges:
        coord_inicial = celdas_rango.coord.split(':')[0]

        letra_columna = coord_inicial[0]  # Obtiene la letra de la columna
        numero_fila_inicial = int(coord_inicial[1:])  # Obtiene el número de la fila inicial

        # Obtiene el valor de la celda fusionada
        valor_celda = hoja[coord_inicial].value
        
        #print(valor_celda)
        codigo_csharp += f'worksheet.Cells["{letra_columna}{numero_fila_inicial}"].Value = "{valor_celda}";\n'

    # Recorre todas las celdas No fusionadas
    for coord_inicial in celdas_no_fusionadas:
        letra_columna = coord_inicial[0]  # Obtiene la letra de la columna
        numero_fila_inicial = int(coord_inicial[1:])  # Obtiene el número de la fila inicial

        # Obtiene el valor de la celda no fusionada
        valor_celda = hoja[coord_inicial].value
        
        codigo_csharp += f'worksheet.Cells["{letra_columna}{numero_fila_inicial}"].Value = "{valor_celda}";\n'

    return codigo_csharp  # Devuelve el código C# generado


def get_column_number(columna):
    num = 0
    for c in columna:
        num = num * 26 + ord(c) - ord('A') + 1
    return num

celda_inicio = 'A11'
celda_fin = 'X13'
rango_celdas = f'{celda_inicio}:{celda_fin}'
celdas_fusionadas = obtener_celdas_fusionadas(hoja)
celdas_no_fusionadas = obtener_celdas_no_fusionadas(hoja, celda_inicio, celda_fin)

print(generar_codigo_csharp_valores(hoja, celdas_fusionadas, celdas_no_fusionadas))


codigo_csharp = generar_codigo_csharp(hoja, celdas_fusionadas, celdas_no_fusionadas)

print(codigo_csharp)

#print(celdas_no_fusionadas)
#print(celdas_fusionadas)


