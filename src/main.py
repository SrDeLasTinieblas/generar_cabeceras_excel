import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.utils import coordinate_to_tuple, get_column_letter
import pyperclip as clipboard

def abrir_archivo_excel(ruta=None):
    if not ruta:
        ruta = input("Ingrese la ruta del archivo Excel (ejemplo: D:\\...\\LibTribInvPermValorizado.xlsx ): ")
    return openpyxl.load_workbook(ruta)

def obtener_rango_celdas(hoja):
    celda_inicio = input("Ingrese la celda de inicio (ejemplo: A13): ")
    celda_fin = input("Ingrese la celda de fin (ejemplo: I14): ")
    return celda_inicio, celda_fin

def obtener_celdas_fusionadas(hoja):
    try:
        celdas_fusionadas = hoja.merged_cells.ranges
        return [str(rango) for rango in celdas_fusionadas]
    except Exception as e:
        print(f"Error al obtener celdas fusionadas: {str(e)}")
        return []

def obtener_celdas_no_fusionadas(hoja, celda_inicio, celda_fin):
    celdas_no_fusionadas = []
    try:
        fila_inicio, columna_inicio = coordinate_to_tuple(celda_inicio)
        fila_fin, columna_fin = coordinate_to_tuple(celda_fin)
        for fila in hoja.iter_rows(min_row=fila_inicio, max_row=fila_fin, min_col=columna_inicio, max_col=columna_fin):
            for celda in fila:
                if not any(celda.coordinate in rango for rango in hoja.merged_cells.ranges):
                    celdas_no_fusionadas.append(celda.coordinate)
        return celdas_no_fusionadas
    except Exception as e:
        print(f"Error al obtener celdas no fusionadas: {str(e)}")
        return []

def obtener_filas_fusionadas(hoja, celda_inicio, celda_fin):
    filas_no_fusionadas = []
    try:
        fila_inicio, _ = coordinate_to_tuple(celda_inicio)
        fila_fin, _ = coordinate_to_tuple(celda_fin)
        for fila in range(fila_inicio, fila_fin + 1):
            if not any((fila, columna) in rango for rango in hoja.merged_cells.ranges):
                filas_no_fusionadas.append(fila)
        return filas_no_fusionadas
    except Exception as e:
        print(f"Error al obtener filas fusionadas: {str(e)}")
        return []

def generar_codigo_csharp(hoja, celdas_fusionadas, celdas_no_fusionadas, filas_fusionadas):
    codigo_csharp = ""
    try:
        # Generar código para celdas fusionadas
        for rango in celdas_fusionadas:
            inicio, fin = rango.split(":")
            estilo_primera_celda = hoja[inicio].font
            nombre_fuente = estilo_primera_celda.name
            tamaño_fuente = int(estilo_primera_celda.size)
            codigo_csharp += f'using (ExcelRange r = worksheet.Cells["{rango}"])\n{{\n'
            codigo_csharp += '    r.Merge = true;\n'
            codigo_csharp += f'    r.Style.Font.SetFromFont(new Font("{nombre_fuente}", {tamaño_fuente}));\n'
            codigo_csharp += '    r.Style.Font.Color.SetColor(Color.Black);\n'
            codigo_csharp += '    r.Style.HorizontalAlignment = ExcelHorizontalAlignment.Center;\n'
            codigo_csharp += '    r.Style.WrapText = true;\n'
            codigo_csharp += '    r.Style.Font.Bold = true;\n'
            codigo_csharp += '    r.Style.Border.Top.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Left.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Right.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Bottom.Style = ExcelBorderStyle.Thin;\n}\n'

        # Generar código para celdas no fusionadas
        for coordenada in celdas_no_fusionadas:
            estilo_celda = hoja[coordenada].font
            nombre_fuente = estilo_celda.name
            tamaño_fuente = int(estilo_celda.size)
            codigo_csharp += f'using (ExcelRange r = worksheet.Cells["{coordenada}"])\n{{\n'
            codigo_csharp += f'    r.Style.Font.SetFromFont(new Font("{nombre_fuente}", {tamaño_fuente}));\n'
            codigo_csharp += '    r.Style.Font.Color.SetColor(Color.Black);\n'
            codigo_csharp += '    r.Style.HorizontalAlignment = ExcelHorizontalAlignment.Center;\n'
            codigo_csharp += '    r.Style.WrapText = true;\n'
            codigo_csharp += '    r.Style.Font.Bold = true;\n'
            codigo_csharp += '    r.Style.Border.Top.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Left.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Right.Style = ExcelBorderStyle.Thin;\n'
            codigo_csharp += '    r.Style.Border.Bottom.Style = ExcelBorderStyle.Thin;\n}\n'

        return codigo_csharp
    except Exception as e:
        print(f"Error al generar código C#: {str(e)}")
        return ""

def generar_codigo_csharp_valores(hoja, celdas_fusionadas, celdas_no_fusionadas):
    codigo_csharp = ""
    for coord in celdas_fusionadas + celdas_no_fusionadas:
        if ":" in coord:
            inicio_rango = coord.split(":")[0]  # Usar solo la primera celda del rango
            valor_celda = hoja[inicio_rango].value or ''
        else:
            valor_celda = hoja[coord].value or ''
        codigo_csharp += f'worksheet.Cells["{coord}"].Value = "{valor_celda}";\n'
    return codigo_csharp

def main():
    workbook = abrir_archivo_excel()
    hoja = workbook.active
    celda_inicio, celda_fin = obtener_rango_celdas(hoja)

    # Obtener celdas y generar código
    celdas_fusionadas = obtener_celdas_fusionadas(hoja)
    celdas_no_fusionadas = obtener_celdas_no_fusionadas(hoja, celda_inicio, celda_fin)
    filas_fusionadas = obtener_filas_fusionadas(hoja, celda_inicio, celda_fin)

    # Generar y copiar código C#
    codigo_csharp = (generar_codigo_csharp_valores(hoja, celdas_fusionadas, celdas_no_fusionadas) +
                     generar_codigo_csharp(hoja, celdas_fusionadas, celdas_no_fusionadas, filas_fusionadas))
    print(codigo_csharp)
    clipboard.copy(codigo_csharp)
    print("Código C# copiado al portapapeles.")

if __name__ == "__main__":
    main()
